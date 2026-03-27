import argparse
import os
import re
import unicodedata
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


TARGET_COLUMNS = ["Event Title", "Start Date", "End Date", "URL", "Location", "Attending"]
COL_WIDTHS = {
    "A": 40,
    "B": 12,
    "C": 13,
    "D": 45,
    "E": 25,
    "F": 15,
}

DATE_FMT = "%m/%d/%Y"
URL_PATTERN = re.compile(r"^https?://", re.IGNORECASE)


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    s = str(value).strip()
    s = s.replace("\u00ae", "(R)")
    s = s.replace("\u2122", "(TM)")
    s = s.replace("\u2019", "'")
    s = s.replace("\u2018", "'")
    s = s.replace("\u201c", '"')
    s = s.replace("\u201d", '"')
    s = s.replace("\u2013", "-")
    s = s.replace("\u2014", "-")
    s = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", s)
    s = unicodedata.normalize("NFKC", s)
    s = "".join(ch for ch in s if unicodedata.category(ch)[0] != "C" or ch in "\n\r\t")
    return s.strip()


def parse_date(value: object):
    if value is None or pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.date()
    raw = str(value).strip()
    if not raw or raw.upper() == "TBD":
        return None

    for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def normalize_date(value: object) -> str:
    d = parse_date(value)
    if d is None:
        raw = clean_text(value)
        return raw if raw else "TBD"
    return d.strftime(DATE_FMT)


def load_input(path: str) -> pd.DataFrame:
    lower = path.lower()
    if lower.endswith(".csv"):
        try:
            return pd.read_csv(path, encoding="utf-8")
        except UnicodeDecodeError:
            try:
                return pd.read_csv(path, encoding="cp1252")
            except UnicodeDecodeError:
                return pd.read_csv(path, encoding="latin1")
    return pd.read_excel(path)


def pick_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    lookup = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        hit = lookup.get(cand.lower())
        if hit is not None:
            return hit
    return None


def build_clean_dataframe(raw_df: pd.DataFrame) -> pd.DataFrame:
    source_map = {
        "Event Title": [
            "Event Title",
            "event_name",
            "Title",
            "Event Name",
            "Name",
        ],
        "Start Date": ["Start Date", "start_date", "Start", "Date Start"],
        "End Date": ["End Date", "end_date", "End", "Date End"],
        "URL": ["URL", "Event URL", "Link", "Website", "url"],
        "Location": ["Location", "Venue", "City", "Country"],
        "Attending": ["Attending"],
    }

    selected = {}
    for target, candidates in source_map.items():
        selected[target] = pick_column(raw_df, candidates)

    rows = []
    for _, row in raw_df.iterrows():
        event_title = clean_text(row[selected["Event Title"]]) if selected["Event Title"] else ""
        if not event_title:
            continue

        start_date = normalize_date(row[selected["Start Date"]]) if selected["Start Date"] else "TBD"
        end_date = normalize_date(row[selected["End Date"]]) if selected["End Date"] else "TBD"
        url = clean_text(row[selected["URL"]]) if selected["URL"] else ""
        location = clean_text(row[selected["Location"]]) if selected["Location"] else "TBD"
        attending = clean_text(row[selected["Attending"]]) if selected["Attending"] else ""

        rows.append(
            {
                "Event Title": event_title,
                "Start Date": start_date,
                "End Date": end_date,
                "URL": url,
                "Location": location if location else "TBD",
                "Attending": attending,
            }
        )

    df = pd.DataFrame(rows, columns=TARGET_COLUMNS)

    # De-duplicate exact event rows to keep output clean.
    df = df.drop_duplicates(subset=["Event Title", "Start Date", "End Date", "URL"], keep="first")

    # Sort by real parsed dates; unknown/invalid dates always go to the bottom.
    df = df.copy()
    df["_start_sort"] = df["Start Date"].map(parse_date)
    df["_end_sort"] = df["End Date"].map(parse_date)
    df["_sort_bucket"] = df["_start_sort"].map(lambda d: 1 if d is None else 0)
    df = df.sort_values(
        by=["_sort_bucket", "_start_sort", "_end_sort", "Event Title"],
        kind="stable",
    )
    df = df.drop(columns=["_start_sort", "_end_sort", "_sort_bucket"]).reset_index(drop=True)

    return df


def find_link_columns(df: pd.DataFrame) -> list[str]:
    return [
        c
        for c in df.columns
        if any(x in str(c).lower() for x in ("linkedin", "url", "link", "website"))
    ]


def style_like_geothermal(path: str, df: pd.DataFrame) -> None:
    wb = load_workbook(path)
    ws = wb.active
    ws.title = "Events"

    link_font = Font(color="0563C1", underline="single")
    header_font = Font(bold=True)
    link_col_names = {str(c).lower() for c in find_link_columns(df)}

    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COL_WIDTHS.get(col_letter, 25)
        ws.cell(row=1, column=col_idx).font = header_font
        ws.cell(row=1, column=col_idx).alignment = Alignment(
            horizontal="center", vertical="top", wrap_text=True
        )

        if str(col_name).lower() in link_col_names:
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = str(cell.value).strip() if cell.value else ""
                if val and URL_PATTERN.match(val):
                    cell.hyperlink = val
                    cell.font = link_font

    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                wrap_text=False, vertical="top"
            )

    max_row, max_col = len(df) + 1, len(df.columns)
    end_letter = get_column_letter(max_col)
    table_ref = f"A1:{end_letter}{max_row}"
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table = Table(displayName="EventsTable", ref=table_ref, tableStyleInfo=style)
    ws.add_table(table)

    wb.save(path)


def default_output_path(input_path: str) -> str:
    base, _ = os.path.splitext(input_path)
    return f"{base}_cleaned.xlsx"


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Create a cleaned Events Excel from CSV/Excel with Geothermal-style formatting."
    )
    parser.add_argument("input", help="Input file path (.csv, .xlsx, .xls)")
    parser.add_argument(
        "--output",
        "-o",
        default=None,
        help="Output Excel path (default: <input>_cleaned.xlsx)",
    )
    args = parser.parse_args()

    input_path = os.path.abspath(args.input)
    output_path = os.path.abspath(args.output) if args.output else default_output_path(input_path)

    raw_df = load_input(input_path)
    clean_df = build_clean_dataframe(raw_df)
    clean_df.to_excel(output_path, index=False)
    style_like_geothermal(output_path, clean_df)

    print(f"Input rows:  {len(raw_df)}")
    print(f"Output rows: {len(clean_df)}")
    print(f"Output file: {output_path}")


if __name__ == "__main__":
    main()
