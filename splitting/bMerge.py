import os
import re
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

mergeDir = "merge"
finalOutputLinks = "CW Speaker List.xlsx"
uniqueCompaniesCsv = "Unique Companies.csv"


def normalizeCompany(name):
    if pd.isna(name) or not isinstance(name, str):
        return ""
    s = str(name).strip().lower()
    return re.sub(r"\s+", " ", s)


def loadCompanyWebsiteMap(baseDir):
    path = os.path.join(baseDir, uniqueCompaniesCsv)
    if not os.path.isfile(path):
        return {}
    try:
        df = pd.read_csv(path)
        if "Company Name" not in df.columns or "Website" not in df.columns:
            return {}
        urlPattern = re.compile(r"^https?://", re.IGNORECASE)
        lookup = {}
        for _, row in df.iterrows():
            company = row.get("Company Name")
            website = row.get("Website")
            if pd.isna(company):
                continue
            norm = normalizeCompany(company)
            if not norm:
                continue
            if pd.notna(website) and urlPattern.search(str(website).strip()):
                lookup[norm] = str(website).strip()
        return lookup
    except Exception:
        return {}


def mergeToFinalCsv():
    baseDir = os.path.dirname(os.path.abspath(__file__))
    mergePath = os.path.join(baseDir, mergeDir)

    if not os.path.isdir(mergePath):
        print("  Merge")
        print("  -----")
        print(f"  x Merge directory not found: {mergePath}")
        return

    pattern = os.path.join(mergePath, "*.csv")
    csvFiles = glob.glob(pattern)

    if not csvFiles:
        print("  Merge")
        print("  -----")
        print(f"  x No CSV files found in {mergePath}")
        return

    def partKey(path):
        m = re.search(r"part_(\d+)", os.path.basename(path), re.IGNORECASE)
        return int(m.group(1)) if m else 0

    csvFiles.sort(key=partKey)

    frames = []
    for path in csvFiles:
        try:
            df = pd.read_csv(path)
            frames.append(df)
        except Exception as e:
            print(f"  x Error reading {path}: {e}")

    if not frames:
        print("  Merge")
        print("  -----")
        print("  x No data to merge.")
        return

    combined = pd.concat(frames, ignore_index=True)
    withLinks = combined

    companyToWebsite = loadCompanyWebsiteMap(baseDir)
    companyCol = None
    for c in combined.columns:
        if str(c).lower() in ("company", "company name"):
            companyCol = c
            break
    if companyCol and companyToWebsite:
        websites = []
        for _, row in combined.iterrows():
            company = row.get(companyCol)
            norm = normalizeCompany(company)
            websites.append(companyToWebsite.get(norm, ""))
        idx = list(combined.columns).index(companyCol) + 1
        combined.insert(idx, "Website", websites)
    elif not companyToWebsite and os.path.isfile(os.path.join(baseDir, uniqueCompaniesCsv)):
        print(f"  Note: {uniqueCompaniesCsv} has no companies with URLs, or column names differ.")
    elif not os.path.isfile(os.path.join(baseDir, uniqueCompaniesCsv)):
        print(f"  Note: {uniqueCompaniesCsv} not found; run scrapeAndAdd.py first.")

    urlCol = None
    for c in combined.columns:
        if "linkedin" in str(c).lower() or "url" in str(c).lower() or "link" in str(c).lower():
            urlCol = c
            break
    urlPattern = re.compile(r"^https?://", re.IGNORECASE)
    if urlCol:
        hasLink = combined[urlCol].astype(str).str.strip().str.startswith(("http://", "https://"), na=False)
        withLinks = combined[hasLink]
        withoutLinks = combined[~hasLink]
        combined = pd.concat([withLinks, withoutLinks], ignore_index=True)

    colWidths = {"Full Name": 30, "Company": 30, "Company Name": 35, "Position": 60, "Website": 35, "LinkedIn": 50}

    def writeExcel(df, path, tableName):
        df.to_excel(path, index=False)
        wb = load_workbook(path)
        ws = wb.active
        linkFont = Font(color="0563C1", underline="single")
        headerFont = Font(bold=True)

        for colIdx, colName in enumerate(df.columns, start=1):
            width = colWidths.get(str(colName), 25)
            ws.column_dimensions[get_column_letter(colIdx)].width = width
            ws.cell(row=1, column=colIdx).font = headerFont
            ws.cell(row=1, column=colIdx).alignment = Alignment(wrap_text=True, vertical="center")

            if (
                "linkedin" in str(colName).lower()
                or "url" in str(colName).lower()
                or "link" in str(colName).lower()
                or "website" in str(colName).lower()
            ):
                for rowIdx in range(2, len(df) + 2):
                    cell = ws.cell(row=rowIdx, column=colIdx)
                    val = str(cell.value).strip() if cell.value else ""
                    if val and urlPattern.match(val):
                        cell.hyperlink = val
                        cell.font = linkFont

        for rowIdx in range(2, len(df) + 2):
            for colIdx in range(1, len(df.columns) + 1):
                ws.cell(row=rowIdx, column=colIdx).alignment = Alignment(wrap_text=False, vertical="top")

        maxRow, maxCol = len(df) + 1, len(df.columns)
        endLetter = get_column_letter(maxCol)
        tableRef = f"A1:{endLetter}{maxRow}"
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table = Table(displayName=tableName, ref=tableRef, tableStyleInfo=style)
        ws.add_table(table)
        wb.save(path)

    if urlCol and len(withLinks) > 0:
        outLinks = os.path.join(baseDir, finalOutputLinks)
        writeExcel(withLinks, outLinks, "SpeakerTableLinks")
        print("  Merge")
        print("  -----")
        print(f"  Rows (links): {len(withLinks)}")
        print(f"  Output:       {outLinks}")


if __name__ == "__main__":
    mergeToFinalCsv()

