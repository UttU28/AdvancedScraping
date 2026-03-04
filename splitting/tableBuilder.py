import os
import re
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

scriptDir = os.path.dirname(os.path.abspath(__file__))
uniqueCompaniesCsv = "Unique Companies.csv"
urlPattern = re.compile(r"^https?://", re.IGNORECASE)

colWidths = {
    "Full Name": 30,
    "Company": 30,
    "Company Name": 35,
    "Position": 60,
    "Website": 35,
    "LinkedIn": 50,
    "Linkedin ": 50,
}


def normalizeCompany(name):
    if pd.isna(name) or not isinstance(name, str):
        return ""
    s = str(name).strip().lower()
    return re.sub(r"\s+", " ", s)


def loadCompanyWebsiteMap():
    path = os.path.join(scriptDir, uniqueCompaniesCsv)
    if not os.path.isfile(path):
        return {}
    try:
        df = pd.read_csv(path)
        if "Company Name" not in df.columns or "Website" not in df.columns:
            return {}
        lookup = {}
        for _, row in df.iterrows():
            company = row.get("Company Name")
            website = row.get("Website")
            if pd.isna(company):
                continue
            norm = normalizeCompany(company)
            if not norm:
                continue
            if pd.notna(website) and urlPattern.search(str(website).strip()):
                lookup[norm] = str(website).strip()
        return lookup
    except Exception:
        return {}


companyColNames = ("Company Name", "Company")


def findCompanyCol(df):
    for col in companyColNames:
        if col in df.columns:
            return col
    for c in df.columns:
        if str(c).strip().lower() in ("company", "company name"):
            return c
    return None


def findLinkColumns(df):
    return [c for c in df.columns if any(x in str(c).lower() for x in ("linkedin", "url", "link", "website"))]


def writeExcel(df, path, tableName="Table1"):
    df.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active
    linkFont = Font(color="0563C1", underline="single")
    headerFont = Font(bold=True)
    linkColNames = {str(c).lower() for c in findLinkColumns(df)}

    for colIdx, colName in enumerate(df.columns, start=1):
        width = colWidths.get(str(colName).strip(), 25)
        ws.column_dimensions[get_column_letter(colIdx)].width = width
        ws.cell(row=1, column=colIdx).font = headerFont
        ws.cell(row=1, column=colIdx).alignment = Alignment(wrap_text=True, vertical="center")

        if str(colName).lower() in linkColNames:
            for rowIdx in range(2, len(df) + 2):
                cell = ws.cell(row=rowIdx, column=colIdx)
                val = str(cell.value).strip() if cell.value else ""
                if val and urlPattern.match(val):
                    cell.hyperlink = val
                    cell.font = linkFont

    for rowIdx in range(2, len(df) + 2):
        for colIdx in range(1, len(df.columns) + 1):
            ws.cell(row=rowIdx, column=colIdx).alignment = Alignment(wrap_text=False, vertical="top")

    maxRow, maxCol = len(df) + 1, len(df.columns)
    endLetter = get_column_letter(maxCol)
    tableRef = f"A1:{endLetter}{maxRow}"
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table = Table(displayName=tableName, ref=tableRef, tableStyleInfo=style)
    ws.add_table(table)
    wb.save(path)


def main():
    if len(sys.argv) < 2:
        print("  Usage: python build_table.py <input.csv|xlsx> [output.xlsx]")
        sys.exit(1)

    inputPath = sys.argv[1]
    if not os.path.isfile(inputPath):
        print(f"  x Input file not found: {inputPath}")
        sys.exit(1)

    outputPath = sys.argv[2] if len(sys.argv) >= 3 else os.path.splitext(inputPath)[0] + "_table.xlsx"

    if inputPath.lower().endswith(".csv"):
        df = pd.read_csv(inputPath)
    else:
        df = pd.read_excel(inputPath)

    companyCol = findCompanyCol(df)
    companyToWebsite = loadCompanyWebsiteMap()

    # Add or update Website column based on company lookup
    if companyCol and companyToWebsite:
        websites = [companyToWebsite.get(normalizeCompany(row.get(companyCol)), "") for _, row in df.iterrows()]

        if "Website" in df.columns:
            # Only fill in missing/empty Website cells, keep existing non-empty values
            df["Website"] = [
                (new if (pd.isna(orig) or str(orig).strip() == "") and new else orig)
                for orig, new in zip(df["Website"], websites)
            ]
        else:
            idx = list(df.columns).index(companyCol) + 1
            df.insert(idx, "Website", websites)

    # Count how many Website cells are non-empty after processing
    filled = 0
    if "Website" in df.columns:
        filled = df["Website"].astype(str).str.strip().ne("").sum()

    linkCol = next((c for c in df.columns if "linkedin" in str(c).lower()), None)
    if linkCol:
        hasLink = df[linkCol].astype(str).str.strip().str.startswith(("http://", "https://"), na=False)
        df = pd.concat([df[hasLink], df[~hasLink]], ignore_index=True)

    writeExcel(df, outputPath, "SpeakerTable")

    print("  Build Table")
    print("  -----------")
    print(f"  Rows:      {len(df)}")
    print(f"  Websites:  {filled}")
    print(f"  Output:    {outputPath}")


if __name__ == "__main__":
    main()
