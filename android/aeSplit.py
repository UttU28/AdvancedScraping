#!/usr/bin/env python3
import math
import os
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


MAX_PEOPLE_PER_FILE = 100
DEFAULT_INPUT_FILE = "Cera Week Utsav.xlsx"
DEFAULT_OUTPUT_DIR = "utsav"

COL_WIDTHS = {
    "Full Name": 30,
    "Company": 35,
    "Position": 60,
    "Website": 35,
    "LinkedIn": 50,
}

EXPECTED_HEADERS = ["Full Name", "Company", "Position", "Website", "LinkedIn"]
HDR_FULL_NAME, HDR_COMPANY, HDR_POSITION, HDR_WEBSITE, HDR_LINKEDIN = EXPECTED_HEADERS
URL_PATTERN = re.compile(r"^https?://", re.IGNORECASE)


def read_rows(input_path):
    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    headers = [ws.cell(row=1, column=i).value for i in range(1, len(EXPECTED_HEADERS) + 1)]
    headers = [str(h).strip() if h is not None else "" for h in headers]
    if headers != EXPECTED_HEADERS:
        raise ValueError(f"Expected headers {EXPECTED_HEADERS}, found {headers}")

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        full_name = ws.cell(row=row_idx, column=1).value
        company = ws.cell(row=row_idx, column=2).value
        position = ws.cell(row=row_idx, column=3).value
        website = ws.cell(row=row_idx, column=4).value
        linkedin = ws.cell(row=row_idx, column=5).value

        if all(v is None or str(v).strip() == "" for v in [full_name, company, position, website, linkedin]):
            continue

        rows.append(
            {
                HDR_FULL_NAME: "" if full_name is None else str(full_name).strip(),
                HDR_COMPANY: "" if company is None else str(company).strip(),
                HDR_POSITION: "" if position is None else str(position).strip(),
                HDR_WEBSITE: "" if website is None else str(website).strip(),
                HDR_LINKEDIN: "" if linkedin is None else str(linkedin).strip(),
            }
        )
    return rows


def write_split_file(rows, output_file, version):
    wb = Workbook()
    ws = wb.active
    ws.title = "People"

    ws.append(EXPECTED_HEADERS)
    header_font = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")

    for col_idx, header in enumerate(EXPECTED_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 25)

    for person in rows:
        ws.append(
            [
                person[HDR_FULL_NAME],
                person[HDR_COMPANY],
                person[HDR_POSITION],
                person[HDR_WEBSITE],
                person[HDR_LINKEDIN],
            ]
        )

    for row_idx in range(2, len(rows) + 2):
        for col_idx in range(1, len(EXPECTED_HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=False, vertical="top")

    for link_col in (4, 5):
        for row_idx in range(2, len(rows) + 2):
            cell = ws.cell(row=row_idx, column=link_col)
            val = str(cell.value).strip() if cell.value else ""
            if val and URL_PATTERN.match(val):
                cell.hyperlink = val
                cell.font = link_font

    if rows:
        max_row = len(rows) + 1
        table_ref = f"A1:E{max_row}"
        table = Table(displayName=f"PeopleTableV{version}", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    wb.save(output_file)


def split_utsav_file(input_file=DEFAULT_INPUT_FILE, output_dir=DEFAULT_OUTPUT_DIR, max_people=MAX_PEOPLE_PER_FILE):
    input_path = Path(input_file)
    if not input_path.is_absolute():
        input_path = Path.cwd() / input_path
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    rows = read_rows(str(input_path))
    total = len(rows)
    if total == 0:
        print("No people rows found in input file.")
        return 0

    output_path = Path(output_dir)
    if not output_path.is_absolute():
        output_path = input_path.parent / output_path
    output_path.mkdir(parents=True, exist_ok=True)

    parts = math.ceil(total / max_people)
    base_name = input_path.stem

    for idx in range(parts):
        start = idx * max_people
        end = min(start + max_people, total)
        chunk = rows[start:end]
        out_file = output_path / f"{base_name} v{idx + 1}.xlsx"
        write_split_file(chunk, str(out_file), idx + 1)

    print(f"Input rows: {total}")
    print(f"Max rows/file: {max_people}")
    print(f"Created {parts} files in: {output_path}")
    return parts


if __name__ == "__main__":
    split_utsav_file()
