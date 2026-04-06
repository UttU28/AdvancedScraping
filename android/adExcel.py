#!/usr/bin/env python3
import json
import os
import re
from difflib import SequenceMatcher

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


COL_WIDTHS = {
    "Full Name": 30,
    "Company": 35,
    "Position": 60,
    "Website": 35,
    "LinkedIn": 50,
}

companyFilters = [
    "University",
    "College",
    "School",
    "Institute",
    "Laboratory",
    "Lab",
    "Software",
    "Consulting",
    "Consultants",
    "Advisors",
    "Advisory",
    "Media",
    "Press",
    "News",
    "Communications",
    "PR",
    "Public Relations",
]
positionFilters = [
    "Analyst",
    "Analysis",
    "Associate",
    "Assistant",
    "Coordinator",
    "Specialist",
    "Representative",
    "Recruiter",
    "Intern",
    "Fellow",
    "Student",
    "PHD",
    "PhD",
    "Professor",
    "Researcher",
    "Human Resource",
    "HR",
    "Journalist",
    "Reporter",
    "Correspondent",
    "Photographer",
    "Social Media",
    "Marketing",
    "Communications",
    "Editorial",
    "Editor",
    "Staff",
    "Sales",
    "Business Development",
    "Account Manager",
    "Customer Success",
    "Partnerships",
    "Manager",
    "Associate Director",
    "Deputy",
    "Engineer",
]
URL_PATTERN = re.compile(r"^https?://", re.IGNORECASE)


def cleanField(value):
    if value is None:
        return ""
    value = str(value).replace("|", " ").strip()
    value = re.sub(r"\.{2,}$", "", value)  # remove OCR-truncated trailing dots
    value = re.sub(r"^[\\/\)\]\}\|,\-:\.;]+", "", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def normalizedKey(name, company, position):
    return (cleanField(name).lower(), cleanField(company).lower(), cleanField(position).lower())


def looseText(value):
    return re.sub(r"[^a-z0-9 ]+", "", cleanField(value).lower()).strip()


def isCloseText(a, b, min_ratio=0.94):
    if not a or not b:
        return False
    if a == b:
        return True
    shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
    if len(shorter) >= 6 and longer.startswith(shorter):
        return True
    return SequenceMatcher(None, a, b).ratio() >= min_ratio


def isNearDuplicate(existing, candidate):
    name_a = looseText(existing["name"])
    name_b = looseText(candidate["name"])
    if not isCloseText(name_a, name_b, min_ratio=0.97):
        return False

    company_a = looseText(existing["company"])
    company_b = looseText(candidate["company"])
    if not isCloseText(company_a, company_b, min_ratio=0.94):
        return False

    position_a = looseText(existing["position"])
    position_b = looseText(candidate["position"])
    return isCloseText(position_a, position_b, min_ratio=0.90)


def cleanAndDeduplicatePeople(data):
    deduped = []
    seen = set()
    by_name = {}

    for person in data:
        name = cleanField(person.get("name", ""))
        company = cleanField(person.get("company", ""))
        position = cleanField(person.get("position", ""))
        if not (name and company and position):
            continue

        key = normalizedKey(name, company, position)
        if key in seen:
            continue

        candidate = {"name": name, "company": company, "position": position}
        name_key = looseText(name)

        near_dupe = False
        for existing in by_name.get(name_key, []):
            if isNearDuplicate(existing, candidate):
                near_dupe = True
                break
        if near_dupe:
            continue

        seen.add(key)
        deduped.append(candidate)
        by_name.setdefault(name_key, []).append(candidate)

    return deduped


def compileKeywordRegex(filters):
    regex_parts = []
    for item in filters:
        token = cleanField(item)
        if not token:
            continue
        # Match flexible spacing while keeping special chars literal.
        token = re.escape(token).replace(r"\ ", r"\s+")
        regex_parts.append(token)
    if not regex_parts:
        return None
    return re.compile(r"(?i)\b(" + "|".join(regex_parts) + r")\b")


def applyFilters(people):
    company_regex = compileKeywordRegex(companyFilters)
    position_regex = compileKeywordRegex(positionFilters)
    filtered = []

    for person in people:
        company = cleanField(person.get("company", ""))
        position = cleanField(person.get("position", ""))

        company_blocked = bool(company_regex and company_regex.search(company))
        position_blocked = bool(position_regex and position_regex.search(position))
        if company_blocked or position_blocked:
            continue
        filtered.append(person)

    return filtered


def writeExcel(people, excelFile):
    wb = Workbook()
    ws = wb.active
    ws.title = "People"

    headers = ["Full Name", "Company", "Position", "Website", "LinkedIn"]
    ws.append(headers)

    header_font = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")
    for col_idx, header in enumerate(headers, start=1):
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.font = header_font
        header_cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 25)

    for person in people:
        ws.append([person["name"], person["company"], person["position"], "", ""])

    for row_idx in range(2, len(people) + 2):
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                wrap_text=False, vertical="top"
            )

    # Make Website and LinkedIn clickable, styled as links.
    for link_col in (4, 5):
        for row_idx in range(2, len(people) + 2):
            cell = ws.cell(row=row_idx, column=link_col)
            val = str(cell.value).strip() if cell.value else ""
            if val and URL_PATTERN.match(val):
                cell.hyperlink = val
                cell.font = link_font

    max_row = len(people) + 1
    table_ref = f"A1:E{max_row}"
    table = Table(displayName="PeopleTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    wb.save(excelFile)


def splitEvenly(items, parts):
    base_size, remainder = divmod(len(items), parts)
    chunks = []
    start = 0
    for i in range(parts):
        size = base_size + (1 if i < remainder else 0)
        end = start + size
        chunks.append(items[start:end])
        start = end
    return chunks


def convertJsonToExcel(jsonFile="people.json", excelFile=None):
    if excelFile is None:
        excelFile = os.path.splitext(jsonFile)[0] + ".xlsx"

    try:
        with open(jsonFile, "r", encoding="utf-8") as f:
            data = json.load(f)

        if not isinstance(data, list):
            print(
                f"Warning: JSON format unexpected. Expected a list of people, got {type(data).__name__}"
            )
            if isinstance(data, dict) and "people" in data:
                data = data["people"]
            else:
                raise ValueError("Cannot process JSON: invalid format")

        deduped = cleanAndDeduplicatePeople(data)
        filtered = applyFilters(deduped)
        writeExcel(filtered, excelFile)

        out_dir = os.path.dirname(excelFile) or "."
        splits = splitEvenly(filtered, 3)
        split_names = ["Utsav", "Sudip", "Parth"]
        split_paths = []
        for person_name, rows in zip(split_names, splits):
            split_file = os.path.join(out_dir, f"Cera Week {person_name}.xlsx")
            writeExcel(rows, split_file)
            split_paths.append(split_file)

        print(f"Successfully converted {jsonFile} to {excelFile}")
        print(
            f"Found {len(data)} people entries, "
            f"wrote {len(filtered)} rows after clean + dedupe + filters"
        )
        print(f"Split files created: {', '.join(split_paths)}")
        return excelFile

    except FileNotFoundError:
        print(f"Error: JSON file not found at {jsonFile}")
        return None
    except json.JSONDecodeError:
        print(f"Error: {jsonFile} is not a valid JSON file")
        return None
    except Exception as e:
        print(f"Error: {str(e)}")
        return None


if __name__ == "__main__":
    convertJsonToExcel()
