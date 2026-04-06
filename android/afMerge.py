#!/usr/bin/env python3
import csv
import glob
import os
import re
from difflib import SequenceMatcher

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


COL_WIDTHS = {
    "Full Name": 30,
    "Company": 35,
    "Position": 60,
    "Website": 35,
    "LinkedIn": 50,
}

HEADERS = ["Full Name", "Company", "Position", "Website", "LinkedIn"]
HDR_FULL_NAME, HDR_COMPANY, HDR_POSITION, HDR_WEBSITE, HDR_LINKEDIN = HEADERS
URL_PATTERN = re.compile(r"^https?://", re.IGNORECASE)


def clean_text(value):
    if value is None:
        return ""
    value = str(value).strip()
    value = re.sub(r"\s+", " ", value)
    return value


def norm_text(value):
    value = clean_text(value).lower()
    value = re.sub(r"[^a-z0-9 ]+", "", value)
    return re.sub(r"\s+", " ", value).strip()


def is_close(a, b, threshold=0.96):
    if not a or not b:
        return False
    if a == b:
        return True
    shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
    if len(shorter) >= 6 and longer.startswith(shorter):
        return True
    return SequenceMatcher(None, a, b).ratio() >= threshold


def read_main_excel(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        full_name = clean_text(ws.cell(row=row_idx, column=1).value)
        company = clean_text(ws.cell(row=row_idx, column=2).value)
        position = clean_text(ws.cell(row=row_idx, column=3).value)
        website = clean_text(ws.cell(row=row_idx, column=4).value)
        linkedin = clean_text(ws.cell(row=row_idx, column=5).value)
        if not (full_name or company or position or website or linkedin):
            continue
        rows.append(
            {
                HDR_FULL_NAME: full_name,
                HDR_COMPANY: company,
                HDR_POSITION: position,
                HDR_WEBSITE: website,
                HDR_LINKEDIN: linkedin,
            }
        )
    return rows


def read_output_csvs(outputs_dir):
    pattern = os.path.join(outputs_dir, "Cera Week Utsav v*xlsx_*.csv")
    files = sorted(glob.glob(pattern))
    rows = []
    for path in files:
        with open(path, "r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            for r in reader:
                rows.append(
                    {
                        HDR_FULL_NAME: clean_text(r.get(HDR_FULL_NAME, "")),
                        HDR_COMPANY: clean_text(r.get(HDR_COMPANY, "")),
                        HDR_POSITION: clean_text(r.get(HDR_POSITION, "")),
                        HDR_WEBSITE: clean_text(r.get(HDR_WEBSITE, "")),
                        HDR_LINKEDIN: clean_text(r.get(HDR_LINKEDIN, "")),
                    }
                )
    return files, rows


def build_lookup(rows):
    by_key = {}
    by_name = {}
    for row in rows:
        name_n = norm_text(row[HDR_FULL_NAME])
        company_n = norm_text(row[HDR_COMPANY])
        if not name_n:
            continue
        key = (name_n, company_n)
        existing = by_key.get(key, {HDR_WEBSITE: "", HDR_LINKEDIN: "", HDR_POSITION: ""})

        # Keep first non-empty value found.
        for field in (HDR_WEBSITE, HDR_LINKEDIN, HDR_POSITION):
            if not existing.get(field) and row.get(field):
                existing[field] = row[field]
        by_key[key] = existing

        by_name.setdefault(name_n, []).append(
            {
                "company_n": company_n,
                HDR_WEBSITE: row.get(HDR_WEBSITE, ""),
                HDR_LINKEDIN: row.get(HDR_LINKEDIN, ""),
                HDR_POSITION: row.get(HDR_POSITION, ""),
            }
        )
    return by_key, by_name


def merge_rows(main_rows, by_key, by_name):
    updated = []
    matched = 0
    for row in main_rows:
        name_n = norm_text(row[HDR_FULL_NAME])
        company_n = norm_text(row[HDR_COMPANY])
        key = (name_n, company_n)
        fill = by_key.get(key)

        if fill is None and name_n in by_name:
            for candidate in by_name[name_n]:
                if is_close(company_n, candidate["company_n"], threshold=0.94):
                    fill = candidate
                    break

        out = dict(row)
        if fill:
            matched += 1
            for field in (HDR_WEBSITE, HDR_LINKEDIN, HDR_POSITION):
                if fill.get(field):
                    out[field] = fill[field]
        updated.append(out)
    return updated, matched


def write_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "People"
    ws.append(HEADERS)

    header_font = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 25)

    for row in rows:
        ws.append([row[h] for h in HEADERS])

    for row_idx in range(2, len(rows) + 2):
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=False, vertical="top")

    for link_col in (4, 5):
        for row_idx in range(2, len(rows) + 2):
            cell = ws.cell(row=row_idx, column=link_col)
            val = str(cell.value).strip() if cell.value else ""
            if val and URL_PATTERN.match(val):
                cell.hyperlink = val
                cell.font = link_font

    if rows:
        table_ref = f"A1:E{len(rows) + 1}"
        table = Table(displayName="PeopleTableMerged", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    wb.save(output_path)


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # Support running from android directly; fall back if script moved under outputs.
    if os.path.isdir(os.path.join(script_dir, "outputs")):
        android_dir = script_dir
    else:
        android_dir = os.path.dirname(script_dir)
    outputs_dir = os.path.join(android_dir, "outputs")
    main_excel = os.path.join(android_dir, "Cera Week Utsav.xlsx")
    output_excel = os.path.join(android_dir, "Cera Week Utsav Updated.xlsx")

    main_rows = read_main_excel(main_excel)
    csv_files, output_rows = read_output_csvs(outputs_dir)
    by_key, by_name = build_lookup(output_rows)
    merged_rows, matched = merge_rows(main_rows, by_key, by_name)
    write_excel(merged_rows, output_excel)

    print(f"Main file: {main_excel}")
    print(f"Outputs directory: {outputs_dir}")
    print(f"Output CSV files merged: {len(csv_files)}")
    print(f"Main rows: {len(main_rows)}")
    print(f"Rows matched with updates: {matched}")
    print(f"Updated file written: {output_excel}")


if __name__ == "__main__":
    main()
