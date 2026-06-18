import argparse
import csv
import os
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


URL_PATTERN = re.compile(r"^https?://", re.IGNORECASE)
FULL_NAME_COL = "Full Name"
NAME_COL = "Name"
COMPANY_COL = "Company"
WEBSITE_COL = "Website"
POSITION_COL = "Position"
LINKEDIN_COL = "LinkedIn"
EMAIL_COL = "Email"
POSITION_TYPO_COL = "Postion"
LINKEDIN_TYPO_COL = "Linkedin "

COL_WIDTHS = {
    FULL_NAME_COL: 30,
    COMPANY_COL: 30,
    "Company Name": 35,
    POSITION_COL: 60,
    WEBSITE_COL: 35,
    LINKEDIN_COL: 50,
    LINKEDIN_TYPO_COL: 50,
    EMAIL_COL: 40,
}


def find_link_columns(df: pd.DataFrame) -> list[str]:
    return [
        c
        for c in df.columns
        if any(x in str(c).lower() for x in ("linkedin", "url", "link", "website"))
    ]


def write_excel_table(df: pd.DataFrame, path: str, table_name: str = "Table1") -> None:
    """Write DataFrame to a formatted Excel file (table style, column widths, link hyperlinks)."""
    df.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active

    link_font = Font(color="0563C1", underline="single")
    header_font = Font(bold=True)
    link_col_names = {str(c).lower() for c in find_link_columns(df)}

    for col_idx, col_name in enumerate(df.columns, start=1):
        width = COL_WIDTHS.get(str(col_name).strip(), 25)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.cell(row=1, column=col_idx).font = header_font
        ws.cell(row=1, column=col_idx).alignment = Alignment(
            wrap_text=True, vertical="center"
        )

        if str(col_name).lower() in link_col_names:
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = str(cell.value).strip() if cell.value else ""
                if val and URL_PATTERN.match(val):
                    cell.hyperlink = val
                    cell.font = link_font

    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                wrap_text=False, vertical="top"
            )

    max_row, max_col = len(df) + 1, len(df.columns)
    end_letter = get_column_letter(max_col)
    table_ref = f"A1:{end_letter}{max_row}"
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table = Table(displayName=table_name, ref=table_ref, tableStyleInfo=style)
    ws.add_table(table)
    wb.save(path)


def normalize_linkedin(value: object) -> str:
    if pd.isna(value):
        return ""
    url = str(value).strip().lower()
    if not url:
        return ""
    url = url.split("?", 1)[0].split("#", 1)[0]
    return url.rstrip("/")


def normalize_company(value: object) -> str:
    if pd.isna(value):
        return ""
    return " ".join(str(value).strip().split()).lower()


def load_csv_with_fallback(path: str) -> pd.DataFrame:
    try:
        return pd.read_csv(path, encoding="utf-8")
    except UnicodeDecodeError:
        try:
            return pd.read_csv(path, encoding="cp1252")
        except UnicodeDecodeError:
            return pd.read_csv(path, encoding="latin1")


def load_table(path: str) -> pd.DataFrame:
    """Load a CSV or Excel file into a DataFrame."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        return pd.read_excel(path)
    if ext == ".csv":
        return load_csv_with_fallback(path)
    raise ValueError(f"Unsupported file type: {path} (use .csv, .xlsx, .xls)")


def filter_rows_by_keywords(df: pd.DataFrame, keywords: list[str]) -> pd.DataFrame:
    """Filter rows by keywords.

    Positive keywords keep matching rows.
    Keywords prefixed with '-' exclude matching rows.
    If only exclude keywords are provided, all rows start included.
    """
    include_kws: list[str] = []
    exclude_kws: list[str] = []
    for raw in keywords:
        token = str(raw).strip().lower()
        if not token:
            continue
        if token.startswith("-") and token[1:].strip():
            exclude_kws.append(token[1:].strip())
        else:
            include_kws.append(token)

    if not include_kws and not exclude_kws:
        return df

    def row_matches(row: pd.Series) -> bool:
        blob_parts: list[str] = []
        for v in row:
            if pd.notna(v):
                blob_parts.append(str(v).lower())
        blob = " ".join(blob_parts)
        include_ok = True if not include_kws else any(k in blob for k in include_kws)
        exclude_hit = any(k in blob for k in exclude_kws)
        return include_ok and not exclude_hit

    mask = df.apply(row_matches, axis=1)
    return df.loc[mask].copy()


def _maybe_keyword_filter(df: pd.DataFrame, keyword_filter: list[str] | None) -> pd.DataFrame:
    if not keyword_filter:
        return df
    return filter_rows_by_keywords(df, keyword_filter)


def _maybe_require_linkedin(df: pd.DataFrame, require_linkedin: bool) -> pd.DataFrame:
    if not require_linkedin:
        return df
    if LINKEDIN_COL not in df.columns:
        return df
    linkedin = df[LINKEDIN_COL].fillna("").astype(str).str.strip()
    return df.loc[linkedin.ne("")].copy()


def _finalize_output_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Keep output columns in a stable order:
    Full Name, Company, Website, Position, LinkedIn, Email, then remaining columns.
    """
    out = df.copy()
    typo_renames: dict[str, str] = {}
    if POSITION_COL not in out.columns and POSITION_TYPO_COL in out.columns:
        typo_renames[POSITION_TYPO_COL] = POSITION_COL
    if LINKEDIN_COL not in out.columns and LINKEDIN_TYPO_COL in out.columns:
        typo_renames[LINKEDIN_TYPO_COL] = LINKEDIN_COL
    if typo_renames:
        out = out.rename(columns=typo_renames)

    if FULL_NAME_COL in out.columns and NAME_COL in out.columns:
        full_name = out[FULL_NAME_COL].fillna("").astype(str).str.strip()
        name = out[NAME_COL].fillna("").astype(str).str.strip()
        out[FULL_NAME_COL] = full_name.where(full_name.ne(""), name)
        out = out.drop(columns=[NAME_COL])
    elif FULL_NAME_COL not in out.columns and NAME_COL in out.columns:
        out = out.rename(columns={NAME_COL: FULL_NAME_COL})

    preferred = [FULL_NAME_COL, COMPANY_COL, WEBSITE_COL, POSITION_COL, LINKEDIN_COL, EMAIL_COL]
    front = [c for c in preferred if c in out.columns]
    rest = [c for c in out.columns if c not in front]
    return out[front + rest]


def dataframe_merge_company(people: pd.DataFrame, companies: pd.DataFrame) -> pd.DataFrame:
    """Join company websites onto people rows by normalized Company name."""
    for col in ("Company",):
        if col not in people.columns:
            raise ValueError(f"Missing '{col}' in base (people) data")
    for col in ("Company", "Website"):
        if col not in companies.columns:
            raise ValueError(f"Missing '{col}' in company lookup data")

    companies = companies.copy()
    companies["_ck"] = companies["Company"].map(normalize_company)
    companies = companies[companies["_ck"].ne("")]
    companies = companies.drop_duplicates(subset=["_ck"], keep="first")
    site_map = dict(zip(companies["_ck"], companies["Website"].fillna("").astype(str).str.strip()))

    out = people.copy()
    out["_ck"] = out["Company"].map(normalize_company)
    out["Website"] = out["_ck"].map(lambda k: site_map.get(k, ""))
    out = out.drop(columns=["_ck"])

    col_order = []
    for want in ("Name", "Company", "Website", "Position", "LinkedIn"):
        if want in out.columns:
            col_order.append(want)
    rest = [c for c in out.columns if c not in col_order]
    return out[col_order + rest]


def dataframe_merge_emails(main_df: pd.DataFrame, emails_df: pd.DataFrame) -> pd.DataFrame:
    """Add Email column by matching LinkedIn URLs."""
    if "LinkedIn" not in main_df.columns:
        raise ValueError("Missing 'LinkedIn' in main data")
    if "LinkedIn" not in emails_df.columns:
        raise ValueError("Missing 'LinkedIn' in email lookup data")

    email_col = "Mails" if "Mails" in emails_df.columns else "Email"
    if email_col not in emails_df.columns:
        raise ValueError("Email file must have 'Email' or 'Mails' column")

    emails_df = emails_df.copy()
    emails_df["_linkedin_key"] = emails_df["LinkedIn"].map(normalize_linkedin)
    emails_df[email_col] = emails_df[email_col].fillna("").astype(str).str.strip()
    emails_df = emails_df[emails_df["_linkedin_key"].ne("") & emails_df[email_col].ne("")]
    emails_df = emails_df.drop_duplicates(subset=["_linkedin_key"], keep="first")
    email_map = dict(zip(emails_df["_linkedin_key"], emails_df[email_col]))

    main_df = main_df.copy()
    main_df["_linkedin_key"] = main_df["LinkedIn"].map(normalize_linkedin)
    main_df["Email"] = main_df["_linkedin_key"].map(email_map).fillna("")
    main_df = main_df.drop(columns=["_linkedin_key"])

    if "LinkedIn" in main_df.columns:
        cols = list(main_df.columns)
        cols.remove("Email")
        linkedin_idx = cols.index("LinkedIn")
        cols.insert(linkedin_idx + 1, "Email")
        main_df = main_df[cols]
    return main_df


def run_format(
    input_path: str,
    output_path: str,
    table_name: str = "Data",
    keyword_filter: list[str] | None = None,
    require_linkedin: bool = False,
) -> int:
    df = load_table(input_path)
    df = _maybe_keyword_filter(df, keyword_filter)
    df = _maybe_require_linkedin(df, require_linkedin)
    df = _finalize_output_columns(df)
    write_excel_table(df, output_path, table_name=table_name)
    return len(df)


def run_merge_company(
    people_path: str,
    companies_path: str,
    output_path: str,
    table_name: str = "EnergyTech",
    keyword_filter: list[str] | None = None,
    require_linkedin: bool = False,
) -> tuple[int, int]:
    """Merge people file with company→website file. Returns (row_count, websites_matched)."""
    people = load_table(people_path)
    companies = load_table(companies_path)
    out = dataframe_merge_company(people, companies)
    out = _maybe_keyword_filter(out, keyword_filter)
    out = _maybe_require_linkedin(out, require_linkedin)
    out = _finalize_output_columns(out)
    write_excel_table(out, output_path, table_name=table_name)
    matched = int(out["Website"].astype(str).str.strip().ne("").sum()) if "Website" in out.columns else 0
    return len(out), int(matched)


def run_merge_emails(
    main_path: str,
    emails_path: str,
    output_path: str,
    table_name: str = "DCDWithEmails",
    keyword_filter: list[str] | None = None,
    require_linkedin: bool = False,
) -> tuple[int, int]:
    """Merge emails into main table by LinkedIn. Returns (row_count, emails_matched)."""
    main_df = load_table(main_path)
    emails_df = load_table(emails_path)
    out = dataframe_merge_emails(main_df, emails_df)
    out = _maybe_keyword_filter(out, keyword_filter)
    out = _maybe_require_linkedin(out, require_linkedin)
    out = _finalize_output_columns(out)
    write_excel_table(out, output_path, table_name=table_name)
    matched = int(out["Email"].astype(str).str.strip().ne("").sum()) if "Email" in out.columns else 0
    return len(out), int(matched)


def _read_csv_rows(path: str) -> list[list[str]]:
    """Read CSV as list of rows, trying common encodings."""
    last_err: OSError | UnicodeDecodeError | None = None
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            with open(path, "r", newline="", encoding=enc) as f:
                return list(csv.reader(f))
        except (UnicodeDecodeError, OSError) as e:
            last_err = e
    if last_err:
        raise last_err
    return []


def split_csv_chunks(input_path: str, output_pattern: str, rows_per_file: int = 40) -> list[str]:
    """Split a CSV into chunk files.

    ``output_pattern`` must contain exactly one ``{}`` placeholder, replaced with the
    1-based chunk index (e.g. ``C:/out/people_{}.csv`` → ``people_1.csv``, ``people_2.csv``, …).
    """
    if rows_per_file < 1:
        raise ValueError("rows_per_file must be at least 1.")
    if output_pattern.count("{}") != 1:
        raise ValueError(
            "Output pattern must contain exactly one '{}' placeholder for the chunk number "
            "(example: C:\\\\data\\\\people_{}.csv)."
        )

    reader = _read_csv_rows(input_path)
    if not reader:
        raise ValueError("CSV is empty or could not be read.")
    header = reader[0]
    data = reader[1:]
    num_chunks = (len(data) + rows_per_file - 1) // rows_per_file
    created: list[str] = []
    for i in range(num_chunks):
        chunk = data[i * rows_per_file : (i + 1) * rows_per_file]
        out_file = os.path.normpath(output_pattern.format(i + 1))
        out_dir = os.path.dirname(out_file)
        if out_dir and not os.path.isdir(out_dir):
            os.makedirs(out_dir, exist_ok=True)
        with open(out_file, "w", newline="", encoding="utf-8") as outfile:
            writer = csv.writer(outfile)
            writer.writerow(header)
            writer.writerows(chunk)
        created.append(out_file)
    return created


def run_merge_full_pipeline(
    base_path: str,
    companies_path: str,
    emails_path: str,
    output_path: str,
    table_name: str = "Merged",
    keyword_filter: list[str] | None = None,
    require_linkedin: bool = False,
) -> tuple[int, int, int]:
    """
    Merge company websites onto base, then merge emails.
    Returns (rows, websites_filled, emails_matched).
    """
    people = load_table(base_path)
    companies = load_table(companies_path)
    emails_df = load_table(emails_path)

    merged = dataframe_merge_company(people, companies)
    merged = dataframe_merge_emails(merged, emails_df)
    merged = _maybe_keyword_filter(merged, keyword_filter)
    merged = _maybe_require_linkedin(merged, require_linkedin)
    merged = _finalize_output_columns(merged)
    web_n = (
        int(merged["Website"].astype(str).str.strip().ne("").sum()) if "Website" in merged.columns else 0
    )
    em_n = int(merged["Email"].astype(str).str.strip().ne("").sum()) if "Email" in merged.columns else 0

    write_excel_table(merged, output_path, table_name=table_name)
    return len(merged), web_n, em_n


def cmd_format(args: argparse.Namespace) -> None:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    in_path = args.input if os.path.isabs(args.input) else os.path.join(script_dir, args.input)
    out_path = args.output if os.path.isabs(args.output) else os.path.join(script_dir, args.output)

    try:
        n = run_format(in_path, out_path, table_name=args.table_name)
    except ValueError as e:
        raise SystemExit(str(e)) from e
    print(f"Rows: {n}")
    print(f"Output: {out_path}")


def cmd_merge_company(args: argparse.Namespace) -> None:
    """Merge people CSV with company→website CSV on Company name, write formatted Excel."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    people_path = args.people if os.path.isabs(args.people) else os.path.join(script_dir, args.people)
    companies_path = (
        args.companies if os.path.isabs(args.companies) else os.path.join(script_dir, args.companies)
    )
    out_path = args.output if os.path.isabs(args.output) else os.path.join(script_dir, args.output)

    try:
        rows, matched = run_merge_company(people_path, companies_path, out_path, table_name=args.table_name)
    except ValueError as e:
        raise SystemExit(str(e)) from e
    companies = load_table(companies_path)
    print(f"People rows:     {rows}")
    print(f"Company rows:    {len(companies)}")
    print(f"Websites filled: {matched}")
    print(f"Output:          {out_path}")


def cmd_merge_emails(args: argparse.Namespace) -> None:
    """Merge email CSV/Excel into main CSV/Excel using LinkedIn URL as key."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    main_path = args.main if os.path.isabs(args.main) else os.path.join(script_dir, args.main)
    emails_path = (
        args.emails if os.path.isabs(args.emails) else os.path.join(script_dir, args.emails)
    )
    output_path = (
        args.output if os.path.isabs(args.output) else os.path.join(script_dir, args.output)
    )

    try:
        rows, matched = run_merge_emails(main_path, emails_path, output_path, table_name=args.table_name)
    except ValueError as e:
        raise SystemExit(str(e)) from e
    emails_df = load_table(emails_path)
    print(f"Main rows:      {rows}")
    print(f"Emails loaded:  {len(emails_df)}")
    print(f"Emails merged:  {matched}")
    print(f"Output file:    {output_path}")


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Format CSV/Excel to styled Excel, merge people with company websites, or merge emails into Excel.",
        epilog="GUI: python dashboard.py",
    )
    sub = p.add_subparsers(dest="command", required=True)

    f = sub.add_parser("format", help="Convert any CSV or Excel to formatted Excel (table + link styling).")
    f.add_argument(
        "-i",
        "--input",
        required=True,
        help="Input .csv, .xlsx, or .xls",
    )
    f.add_argument(
        "-o",
        "--output",
        required=True,
        help="Output .xlsx path",
    )
    f.add_argument(
        "--table-name",
        default="Data",
        help="Excel table name (default: Data)",
    )
    f.set_defaults(_run=cmd_format)

    s = sub.add_parser(
        "merge-company",
        help="Merge people CSV with Company+Website CSV (join on company name), output Excel.",
    )
    s.add_argument(
        "--people",
        default="EnergyTechSummit.csv",
        help="CSV with Name, Company, Position, LinkedIn (default: EnergyTechSummit.csv)",
    )
    s.add_argument(
        "--companies",
        default="cssv.csv",
        help="CSV with Company, Website (default: cssv.csv)",
    )
    s.add_argument(
        "-o",
        "--output",
        default="EnergyTechSummit_merged.xlsx",
        help="Output Excel path",
    )
    s.add_argument(
        "--table-name",
        default="EnergyTech",
        help="Excel table name",
    )
    s.set_defaults(_run=cmd_merge_company)

    e = sub.add_parser(
        "merge-emails",
        help="Merge email CSV into main Excel using LinkedIn URL as key (DCD-style).",
    )
    e.add_argument(
        "--main",
        default="DCD App Mar 20.xlsx",
        help="Main Excel file path",
    )
    e.add_argument(
        "--emails",
        default="DCDEmail.csv",
        help="CSV with LinkedIn and Email (or Mails)",
    )
    e.add_argument(
        "-o",
        "--output",
        default="DCD App Mar 20_with_emails.xlsx",
        help="Output Excel path",
    )
    e.add_argument(
        "--table-name",
        default="DCDWithEmails",
        help="Excel table name",
    )
    e.set_defaults(_run=cmd_merge_emails)

    return p


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    args._run(args)


if __name__ == "__main__":
    main()
